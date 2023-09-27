import openpyxl
from openpyxl.styles import PatternFill

file = 'Book1.xlsx'
wb = openpyxl.load_workbook(file)
ws = wb.active

ministry_column = ws['A']
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')

for cell in ministry_column:
    if cell.value == 'Rahul':
        for row in ws.iter_rows(min_row=1, min_col=cell.column, max_col=cell.column):
            for cell_in_col in row:
                cell_in_col.fill = yellow_fill

wb.save('Book2.xlsx')
